import openpyxl
from datetime import datetime
from app import df

# Carregar a planilha
caminho_da_planilha = 'matriculas.xlsx'
book = openpyxl.load_workbook(df)
sheet = book.active

# Converter as datas para datetime e coletar as linhas
linhas_com_datas = []
for row in sheet.iter_rows(min_row=2, values_only=True):  # Ignora o cabeçalho, e segue em diante até o final da planilha.
    data = datetime.strptime(row[0], '%d/%m/%Y')  # Ajuste o formato da data conforme necessário, padrão brasileiro.
    linhas_com_datas.append((data, row))

# Ordenar as linhas pela data
linhas_com_datas.sort(key=lambda x: x[0])

# Reescrever as linhas ordenadas na planilha
for i, (_, row) in enumerate(linhas_com_datas, start=2):  # Começa na segunda linha para preservar o cabeçalho.
    for j, value in enumerate(row, start=1):  # Começa em "1" para corresponder aos índices do Excel.
        sheet.cell(row=i, column=j, value=value)

# Salvar as alterações em um novo arquivo para preservar o original.
book.save('matriculas-piaf.xlsx')

# Foram mais de 5 meses para fazer esse código funcionar...
# Tentei em typescript, javascript, cogitei em até usar julia, e nada.
# Mas, finalmente, consegui fazer o código funcionar em python, assim o projeto andou,
# Já que python é uma linguagem que tem uma biblioteca quase que nativa para planilhas.
# E às 23:05 do dia 24 de julho de 2024 encerro a parte mais importante, o funcionamento.
# 
# I never knew, I never thought
# Such bliss as this could fill me with a love divine
# I'm afraid I'll wake and find it was only in my mind
# Was it a dream, or are you really mine?
#
# ...